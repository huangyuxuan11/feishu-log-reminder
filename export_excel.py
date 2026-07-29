#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
客户经理日志 - 每日Excel导出 (设计运行于 GitHub Actions，云端执行)

逻辑：
  1. 用飞书自建应用的 app_id/app_secret 换取 tenant_access_token
  2. 读取「客户经理日志统计」多维表格，筛选当天(填写日期==今天)的记录
  3. 把当天所有填报记录导出为 Excel（列：填写日期/客户经理姓名/工作时间1/工作内容1/工作时间2/工作内容2…，时段成对交叉）
  4. 通过飞书自建应用，把 Excel 作为文件私信发给你（亦非fan）

环境变量(放 GitHub Secrets)：
  FEISHU_APP_ID          飞书自建应用 app_id
  FEISHU_APP_SECRET      飞书自建应用 app_secret
  FEISHU_USER_OPEN_ID    接收人飞书 open_id（亦非fan）
  FEISHU_USER_ID         可选，留空即可

注意：发文件需要应用具备 im:resource 权限（文件上传）+ im:message + im:message:send_as_bot。
"""

import os
import io
import re
import json
import time
import datetime
import requests

# ---------- 配置：客户经理日志统计 多维表格 ----------
LOG_APP = "OteIbzKYha9jKKsprhwcKovNnCh"
LOG_TABLE = "tblAQ9ZGEkAmBgTT"

FEISHU_BASE = "https://open.feishu.cn/open-apis"
BJT = datetime.timezone(datetime.timedelta(hours=8))


def get_tenant_token(app_id, app_secret):
    url = f"{FEISHU_BASE}/auth/v3/tenant_access_token/internal"
    r = requests.post(url, json={"app_id": app_id, "app_secret": app_secret}, timeout=15)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"获取 tenant_access_token 失败: {data}")
    return data["tenant_access_token"]


def _norm(v):
    if v is None:
        return ""
    if isinstance(v, list):
        return _norm(v[0]) if v else ""
    if isinstance(v, dict):
        for k in ("date", "start"):
            if k in v:
                return str(v[k])
        return str(v)
    return str(v)


_DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _date_only(v):
    """从飞书日期字段值中提取 YYYY-MM-DD。

    飞书日期时间型字段经开放 API 返回格式不固定：
      - 整数/浮点毫秒时间戳（如 1785168000000）→ 按 BJT 时区转 datetime
      - 字符串 '2026-07-28 00:00:00' / '2026-07-28' → 正则提取日期部分
      - dict {'date': ...} / list [...] → 取内部值
    统一返回 'YYYY-MM-DD'，用于与 today 字符串比较。
    """
    # dict / list 展开
    if isinstance(v, dict):
        for k in ("date", "start"):
            if k in v:
                v = v[k]
                break
    if isinstance(v, list):
        v = v[0] if v else ""
    # 整数/浮点 → 视为毫秒时间戳
    if isinstance(v, bool):
        pass
    elif isinstance(v, (int, float)):
        try:
            return datetime.datetime.fromtimestamp(v / 1000, tz=BJT).strftime("%Y-%m-%d")
        except (ValueError, OSError, OverflowError):
            return str(v)
    s = _norm(v) if not isinstance(v, str) else v
    if isinstance(s, str):
        s = s.strip()
        m = _DATE_RE.search(s)
        return m.group(1) if m else s
    return str(v)


def list_all_records(app_token, table_id, token, page_size=100):
    url = f"{FEISHU_BASE}/bitable/v1/apps/{app_token}/tables/{table_id}/records"
    out, page_token = [], ""
    while True:
        params = {"page_size": page_size}
        if page_token:
            params["page_token"] = page_token
        r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, params=params, timeout=20)
        data = r.json()
        if data.get("code") != 0:
            raise RuntimeError(f"读取记录失败: {data}")
        out.extend(data.get("data", {}).get("items", []))
        if not data.get("data", {}).get("has_more") or not data.get("data", {}).get("page_token"):
            break
        page_token = data["data"]["page_token"]
    return out


# ---------- 去重控制表（避免飞书自动化 与 GitHub 自带定时器 双触发导致重复发文件） ----------
CONTROL_APP = "H0qGbDCJiaMXQJss7xOcvTYEnsh"   # 日志提醒去重控制
CONTROL_TABLE = "tbleEVzbXaicZ19E"

# 北京时间各触发档位（与飞书自动化 / GitHub schedule 一致）
SLOTS = [(2000, 20, 0), (2030, 20, 30), (2045, 20, 45), (2130, 21, 30)]


def get_slot(now):
    """返回当前时间最接近的档位 key（±20 分钟内），否则 None（如手动测试）。"""
    cur = now.hour * 60 + now.minute
    best, best_diff = None, 999
    for key, h, m in SLOTS:
        diff = abs(cur - (h * 60 + m))
        if diff < best_diff:
            best, best_diff = key, diff
    return best if best_diff <= 20 else None


def _create_record(app_token, table_id, fields, token):
    url = f"{FEISHU_BASE}/bitable/v1/apps/{app_token}/tables/{table_id}/records"
    r = requests.post(url, headers={"Authorization": f"Bearer {token}"}, json={"fields": fields}, timeout=20)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"创建记录失败: {data}")
    return data.get("data", {}).get("record")


def _update_record(app_token, table_id, record_id, fields, token):
    url = f"{FEISHU_BASE}/bitable/v1/apps/{app_token}/tables/{table_id}/records/{record_id}"
    r = requests.post(url, headers={"Authorization": f"Bearer {token}"}, json={"fields": fields}, timeout=20)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"更新记录失败: {data}")
    return data


def _find_control_record(records, task):
    for it in records:
        raw = _norm(it.get("fields", {}).get("文本"))
        if raw.startswith(f"{task}|"):
            return it, raw
    return None, ""


def read_fired_slots(token, task, today):
    """读取该任务今天已发送的档位集合；异常时返回空集（fail-open，不阻断发送）。"""
    try:
        records = list_all_records(CONTROL_APP, CONTROL_TABLE, token)
        it, raw = _find_control_record(records, task)
        if it and raw:
            parts = raw.split("|")
            if len(parts) >= 3 and parts[1] == today:
                return set(p for p in parts[2].split(",") if p)
    except Exception as e:
        print(f"[dedup] 读取控制表失败(忽略，继续发送): {e}")
    return set()


def mark_slot_fired(token, task, today, slot):
    """标记该任务今天某档位已发送；异常时仅记录（fail-open）。"""
    try:
        records = list_all_records(CONTROL_APP, CONTROL_TABLE, token)
        it, raw = _find_control_record(records, task)
        if it:
            rec_id = it.get("record_id") or it.get("id")
            parts = raw.split("|")
            if len(parts) >= 3 and parts[1] == today:
                slots = set(p for p in parts[2].split(",") if p)
                slots.add(str(slot))
                _update_record(CONTROL_APP, CONTROL_TABLE, rec_id,
                               {"文本": f"{task}|{today}|{','.join(sorted(slots))}"}, token)
            else:
                _update_record(CONTROL_APP, CONTROL_TABLE, rec_id,
                               {"文本": f"{task}|{today}|{slot}"}, token)
        else:
            _create_record(CONTROL_APP, CONTROL_TABLE, {"文本": f"{task}|{today}|{slot}"}, token)
    except Exception as e:
        print(f"[dedup] 写入控制表失败(忽略): {e}")


def get_today_records(token, today):
    """当天填报记录列表。

    飞书日期时间型字段的服务端筛选(isGreater/isLess)对 value 格式要求苛刻
    且文档不明确，多次尝试(日期字符串/毫秒时间戳)均报 InvalidFilter。
    故采用稳定策略：全量拉取 + 本地 _date_only 过滤。
    数据量约 1300+ 条，API 翻页拉取耗时可控（<5秒）。
    """
    items = list_all_records(LOG_APP, LOG_TABLE, token)
    print(f"[debug] 全量拉取总记录数: {len(items)}")
    matched = [it for it in items if _date_only(it.get("fields", {}).get("填写日期")) == today]
    print(f"[debug] 匹配今天({today})的记录数: {len(matched)}")
    return matched


def build_excel(records, today_label):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # 找出最大的时段序号 N（字段形如 工作时间N / 工作内容N）
    max_n = 0
    for it in records:
        for k in it.get("fields", {}):
            m = re.match(r"^工作时间(\d+)$", k)
            if m:
                max_n = max(max_n, int(m.group(1)))
                continue
            m = re.match(r"^工作内容(\d+)$", k)
            if m:
                max_n = max(max_n, int(m.group(1)))

    # 多列、成对交叉：填写日期/客户经理姓名/工作时间1/工作内容1/工作时间2/工作内容2/...
    headers = ["填写日期", "客户经理姓名"]
    for n in range(1, max_n + 1):
        headers.append(f"工作时间{n}")
        headers.append(f"工作内容{n}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户经理日志"

    # 表头样式
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 数据行
    for r, it in enumerate(records, 2):
        f = it.get("fields", {})
        row_vals = [_date_only(f.get("填写日期")), _norm(f.get("客户经理姓名"))]
        for n in range(1, max_n + 1):
            row_vals.append(_norm(f.get(f"工作时间{n}")))
            row_vals.append(_norm(f.get(f"工作内容{n}")))
        for c, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # 列宽
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 12
    for n in range(1, max_n + 1):
        tcol = 3 + (n - 1) * 2
        ccol = tcol + 1
        ws.column_dimensions[openpyxl.utils.get_column_letter(tcol)].width = 16
        ws.column_dimensions[openpyxl.utils.get_column_letter(ccol)].width = 40

    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def upload_file(token, file_bytes, file_name):
    url = f"{FEISHU_BASE}/im/v1/files"
    files = {
        "file": (file_name, file_bytes,
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    }
    data = {"file_name": file_name, "file_type": "xlsx"}
    r = requests.post(url, headers={"Authorization": f"Bearer {token}"},
                      files=files, data=data, timeout=30)
    resp = r.json()
    if resp.get("code") != 0:
        raise RuntimeError(f"文件上传失败: {resp}")
    return resp["data"]["file_key"]


def send_file_message(token, open_id, file_key):
    url = f"{FEISHU_BASE}/im/v1/messages?receive_id_type=open_id"
    body = {
        "receive_id": open_id,
        "msg_type": "file",
        "content": json.dumps({"file_key": file_key}),
    }
    r = requests.post(url, headers={"Authorization": f"Bearer {token}"}, json=body, timeout=20)
    return r.json()


def send_text_message(token, open_id, text):
    url = f"{FEISHU_BASE}/im/v1/messages?receive_id_type=open_id"
    body = {
        "receive_id": open_id,
        "msg_type": "text",
        "content": json.dumps({"text": text}),
    }
    r = requests.post(url, headers={"Authorization": f"Bearer {token}"}, json=body, timeout=20)
    return r.json()


def main():
    app_id = os.environ.get("FEISHU_APP_ID")
    app_secret = os.environ.get("FEISHU_APP_SECRET")
    open_id = os.environ.get("FEISHU_USER_OPEN_ID")
    if not app_id or not app_secret or not open_id:
        raise SystemExit("缺少环境变量 FEISHU_APP_ID / FEISHU_APP_SECRET / FEISHU_USER_OPEN_ID")

    now = datetime.datetime.now(BJT)
    today = now.strftime("%Y-%m-%d")
    today_label = now.strftime("%Y年%m月%d日")

    token = get_tenant_token(app_id, app_secret)

    # 去重：若本档位今日已由另一触发器（飞书自动化 / GitHub 自带定时器）导出过，跳过防重复
    slot = get_slot(now)
    if slot is not None:
        if str(slot) in read_fired_slots(token, "export", today):
            print(f"[dedup] 档位 {slot} 今日已导出，跳过（防重复）")
            return

    records = get_today_records(token, today)
    print(f"[info] 当天({today})填报记录数: {len(records)}")

    if not records:
        msg = f"📭 {today_label} 暂无客户经理填写日志，未导出Excel。"
        print(msg)
        resp = send_text_message(token, open_id, msg)
        print("text 返回:", resp)
        if slot is not None:
            mark_slot_fired(token, "export", today, slot)
        return

    buf = build_excel(records, today_label)
    file_name = f"客户经理日志_{today}.xlsx"
    file_key = upload_file(token, buf.getvalue(), file_name)
    print(f"[info] 文件已上传 file_key={file_key}")

    # 先发一句说明，再发文件
    intro = f"📊 {today_label} 客户经理日志导出（共 {len(records)} 条填报），见附件Excel。"
    send_text_message(token, open_id, intro)
    resp = send_file_message(token, open_id, file_key)
    print("file 返回:", resp)
    print(f"✅ 已发送 {file_name}")
    if slot is not None:
        mark_slot_fired(token, "export", today, slot)


if __name__ == "__main__":
    main()
